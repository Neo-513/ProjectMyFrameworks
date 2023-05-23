import csv
import json
import openpyxl
import pickle
import xlrd
import xlwt

DESKTOP = "C:/Users/Dell/Desktop/"  # 桌面地址


class MyFileIO:
	@staticmethod
	def read(path, encoding):
		with open(path, mode="r", encoding=encoding) as file:
			return file.read()

	@staticmethod
	def write(path, datas, encoding):
		with open(path, mode="w", encoding=encoding) as file:
			file.write(datas)


class MyCsv(MyFileIO):
	@staticmethod
	def read(path, encoding):
		with open(path, mode="r", encoding=encoding) as file:
			return [[r for r in reader] for reader in csv.reader(file)]

	@staticmethod
	def write(path, datas, encoding):
		with open(path, mode="w", encoding=encoding, newline="") as file:
			csv.writer(file).writerows(datas)


class MyJson(MyFileIO):
	@staticmethod
	def read(path, encoding):
		with open(path, mode="r", encoding=encoding) as file:
			return json.load(file)

	@staticmethod
	def write(path, datas, encoding):
		with open(path, mode="w", encoding=encoding) as file:
			json.dump(datas, file)


class MyPickle(MyFileIO):
	@staticmethod
	def read(path, encoding):
		with open(path, mode="rb") as file:
			return pickle.load(file)

	@staticmethod
	def write(path, datas, encoding):
		with open(path, mode="wb") as file:
			pickle.dump(datas, file)


class MyXls(MyFileIO):
	@staticmethod
	def read(path, encoding):
		workbook = xlrd.open_workbook(path)  # 获取工作簿
		worksheet = workbook.sheet_by_index(0)  # 获取首个工作表
		return [[d.value if d.value is not None else "" for d in data] for data in worksheet.get_rows()]

	@staticmethod
	def write(path, datas, encoding):
		workbook = xlwt.Workbook()  # 创建工作簿
		worksheet = workbook.add_sheet("Sheet")  # 创建工作表

		for i, data in enumerate(datas):
			for j, d in enumerate(data):
				worksheet.write(i, j, d)
		workbook.save(path)  # 保存工作簿


class MyXlsx(MyFileIO):
	@staticmethod
	def read(path, encoding):
		workbook = openpyxl.load_workbook(path, read_only=True)  # 获取工作簿（只读模式）
		worksheet = workbook.active  # 获取当前工作表

		datas = [[d.value if d.value is not None else "" for d in data] for data in worksheet.rows]
		workbook.close()  # 关闭工作簿
		return datas

	@staticmethod
	def write(path, datas, encoding):
		workbook = openpyxl.Workbook(write_only=True)  # 创建工作簿（只写模式）
		worksheet = workbook.create_sheet()  # 创建工作表

		for data in datas:
			worksheet.append(data)
		workbook.save(path)  # 保存工作簿
		workbook.close()  # 关闭工作簿


FILE_TYPES = {
	"csv": MyCsv,
	"json": MyJson,
	"pkl": MyPickle,
	"xls": MyXls,
	"xlsx": MyXlsx
}  # 文件类型


def read(path, encoding="utf-8"):
	"""
	读取文件
	:param path: 文件路径
	:param encoding: 文件编码
	:return: 文件内容
	"""
	file_type = path.split(".")[-1]  # 文件类型
	if file_type in FILE_TYPES:
		return FILE_TYPES[file_type].read(path, encoding)
	else:
		return MyFileIO.read(path, encoding)  # 通用读写类


def write(path, datas, encoding="utf-8"):
	"""
	写入文件
	:param path: 文件路径
	:param datas: 文件内容
	:param encoding: 文件编码
	"""
	file_type = path.split(".")[-1]  # 文件类型
	if file_type in FILE_TYPES:
		FILE_TYPES[file_type].write(path, datas, encoding)
	else:
		MyFileIO.write(path, datas, encoding)  # 通用读写类
